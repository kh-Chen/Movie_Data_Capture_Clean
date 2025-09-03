import random
import openpyxl
import sys
import os
import unicodedata
import traceback

def sample_keep_order(lst, n):
    if n <= 0:
        return []
    n = min(n, len(lst))  # 确保n不超过列表长度
    indices = sorted(random.sample(range(len(lst)), n))
    return [lst[i] for i in indices]

def get_display_width(text):
    width = 0
    for char in str(text):
        char_width = 2 if is_wide_character(char) else 1
        width += char_width
    return width

def get_terminal_width():
    try:
        return os.get_terminal_size().columns - 4
    except OSError:
        return 120

def is_wide_character(char):
    return unicodedata.east_asian_width(char) in ('F', 'W')

def pad_to_width_optimized(text, width):
    text = str(text)
    ellipsis = '…'
    ellipsis_width = get_display_width(ellipsis)
    
    current_width = 0
    char_list = []
    for char in text:
        char_width = 2 if is_wide_character(char) else 1
        if current_width + char_width > width:
            break
        current_width += char_width
        char_list.append(char)
    
    if len(char_list) != len(text):
        if current_width == width:
            c = char_list[-1]
            char_list = char_list[:-1]
            if is_wide_character(c):
                char_list.append(' ')

    result = ''.join(char_list)
    if result != text:
        result += ellipsis
        current_width += ellipsis_width
    
    return result + ' ' * (width - current_width)

def countlength(needed_widths,col_index,target_len):
    col_values = [row[col_index] for row in needed_widths]
    t = 0
    for val in col_values:
        if val > target_len:
            t += (val-target_len)
    return t


def calculate_column_widths_new(all_width, needed_widths, min_width=5):
    if not needed_widths:
        return []
    
    if not needed_widths[0]:
        return []
    
    n_cols = len(needed_widths[0])
    max_widths = []
    final_widths = []
    for col_idx in range(n_cols):
        col_values = [row[col_idx] for row in needed_widths]
        col_max = max(col_values)
        max_widths.append(col_max)
        final_widths.append(col_max)
    
    max_width = sum(max_widths)
    if all_width >= max_width:
        return max_widths
    
    min_total = n_cols * min_width
    if min_total > all_width:
        raise ValueError(f"即使所有列都使用最小宽度({min_width})，总宽度({min_total})仍超过屏幕宽度({all_width})")
    
    aa = max_width-all_width
    for i in range(aa):
        lens = [countlength(needed_widths,col_idx,final_widths[col_idx]-1) for col_idx in range(n_cols)]
        
        index = 0
        minlen = 99999
        for idx,length in enumerate(lens):
            if final_widths[idx] <= min_width:
                continue
            if length<minlen:
                index = idx
                minlen = length
            elif length == minlen:
                if max_widths[idx] > max_widths[index]:
                    index = idx
                
        final_widths[index] -= 1
        for row in needed_widths:
            if row[index]>final_widths[index]:
               row[index] -= 1

    return final_widths

def get_row_idx(max_row:int, start:int, limit:int):
    printrowindexs = list(range(2, max_row + 1))
    if limit < max_row:
        if start < 0:
            printrowindexs = random.sample(printrowindexs, limit)
        else:
            if len(printrowindexs) < start+limit:
                printrowindexs = printrowindexs[:limit]
            else:
                printrowindexs = printrowindexs[start:start+limit]
    return printrowindexs

def print_count(max_row=0,search_row=0,show_row=0):
    if search_row < 0:
        print(f"总计 {max_row} 行，显示 {show_row} 行 ")
    else:
        print(f"总计 {max_row} 行，检索到 {search_row} 行，显示 {show_row} 行 ")

def print_data(print_data_widths:slice, while_print_data:slice):
    n_cols = len(while_print_data[0])
    
    terminal_width = get_terminal_width()
    spitcharlen = (n_cols-1)*3+3

    col_display_widths_print = calculate_column_widths_new(terminal_width-spitcharlen,print_data_widths)
    
    separator = "-" * (sum(col_display_widths_print)+spitcharlen+1)
    for idx, print_data_line in enumerate(while_print_data):
        if idx == 0 and n_cols!=1:
            print(separator)
            header = " | ".join(pad_to_width_optimized(print_data_line[i],col_display_widths_print[i]) for i in range(n_cols))
            print("| "+header+" |")
            print(separator)
        else:
            if n_cols!=1:
                line = " | ".join(pad_to_width_optimized(print_data_line[i],col_display_widths_print[i]) for i in range(n_cols))
                print("| "+line+" |")
            else:
                print(pad_to_width_optimized(print_data_line[0],col_display_widths_print[0]))
    
    if n_cols != 1:
        print(separator)

def handle_limit(start=0, limit=20, use_random=False,content=[]):
    total_rows = len(content)
    start-=2
    start = max(0, min(start, total_rows))  
    
    if use_random:  
        remaining = content[start:]
        if limit == -1:  
            content = remaining
        else:
            if len(remaining) > limit:
                content = sample_keep_order(remaining, limit)
            else:
                content = remaining
    else:
        if limit == -1: 
            content = content[start:]
        else:
            end_idx = start + limit
            content = content[start:end_idx]
    return content

def read_xlsx(file_path,search_str=None, search_cols=None, cols=[]):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    
    if len(cols) == 0:
        cols = [col_idx+1 for col_idx in range(max_column)]
    
    all_data = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if search_str and row_idx != 1:
            search_str = search_str.lower()
            flag = False
            for col_idx, cell_value in enumerate(row):
                if search_cols and (col_idx + 1) not in search_cols:
                    continue
                if search_str in str(cell_value).lower():
                    flag=True
                    break
            if not flag:
                continue

        data_line = []
        for col_idx in cols:
            if col_idx == 0:  
                cell_data = str(row_idx) 
                if row_idx == 1:  
                    cell_data = "#"
            else:
                cell_data = row[col_idx-1]
            if cell_data is None:
                cell_data = ""
            cell_data = str(cell_data).strip()
            if cell_data == "" and row_idx != 1: 
                cell_data = str(row_idx)
            data_line.append(cell_data)
        all_data.append(data_line)
    return max_row,all_data
    

def action(file_path, cols=[], start=0, limit=20, use_random=False, search_str=None, search_cols=None):
    try:
        max_row,all_data = read_xlsx(file_path,search_str,search_cols,cols)
        
        content = handle_limit(start,limit,use_random,all_data[1:])
        
        while_print_data = [all_data[0]]
        print_data_widths = []
        for data_line in content:
            print_data_line = []
            print_data_width_line = []
            for cell_value in data_line:
                print_data_line.append(cell_value)
                print_data_width_line.append(get_display_width(cell_value))
            while_print_data.append(print_data_line)
            print_data_widths.append(print_data_width_line)
        
        if content:
            print_data(print_data_widths, while_print_data)
            print_count(max_row,len(all_data)-1 if search_str else -1, len(while_print_data)-1)
        else:
            print("没有数据可显示")
        
    except FileNotFoundError:
        print(f"错误: 文件 '{file_path}' 未找到")
        sys.exit(1)
    except Exception as e:
        print(f"读取文件时出错: {e}")
        traceback.print_exc()
        sys.exit(1)


import argparse

def parse_arguments():
    parser = argparse.ArgumentParser(description='读取并显示Excel文件内容')
    parser.add_argument('file_path', help='Excel文件路径')
    parser.add_argument('columns', nargs='?', default='', help='要显示的列索引（逗号分隔，如:1,2,3)，可选，默认为全部列')
    parser.add_argument('--start', type=int, default=0, help='起始行号，默认为0')
    parser.add_argument('--limit', default='10', help='输出行数，默认为10，可设为"all"输出全部数据')
    parser.add_argument('--random', action='store_true', help='启用随机采样')
    parser.add_argument('--search', help='关键字检索内容（可选）')
    parser.add_argument('--search-columns', help='检索列索引（逗号分隔，可选，默认所有列）')
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_arguments()
    
    file_path = args.file_path
    colnames = args.columns
    
    cols = []
    if colnames: 
        cols = [int(c) for c in colnames.split(",") if c.isdigit()]
    
    limit = args.limit
    if limit.lower() == 'all':
        limit = -1 
    else:
        try:
            limit = int(limit)
        except ValueError:
            print("错误：limit参数必须是整数或'all'")
            sys.exit(1)
    
    # 处理检索列参数
    search_cols = None
    if args.search_columns:
        search_cols = [int(c) for c in args.search_columns.split(",") if c.isdigit()]
    
    action(file_path, cols, args.start, limit, args.random, args.search, search_cols)


    