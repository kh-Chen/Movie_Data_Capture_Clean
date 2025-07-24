import random
import openpyxl
import sys
import unicodedata

chararr = ['…','●']
def get_display_width(text):
    """计算字符串的显示宽度（考虑宽窄字符）"""
    width = 0
    for char in text:
        char_width = 2 if char not in chararr and unicodedata.east_asian_width(char) in ('F', 'W', 'A') else 1
        width += char_width
    return width

def format_cell(cell_value, column_title):
    """根据列标题格式化单元格内容"""
    if cell_value is None:
        return ""
    
    text = str(cell_value)
    
    # 根据列标题应用不同的字符限制
    if column_title == "标题":
        max_chars = 46
    elif column_title == "演员":
        max_chars = 10
    else:
        max_chars = 1000  # 其他列不限制
    
    # 截断超长文本（考虑字符数量）
    if len(text) > max_chars:
        return text[:max_chars-3] + "..."
    return text

def pad_to_width(text, width):
    """填充文本到指定显示宽度（考虑宽窄字符）"""
    current_width = get_display_width(text)
    if current_width >= width:
        return text
    
    # 计算需要添加的空格数量
    padding = width - current_width
    return text + ' ' * padding

def read_xlsx(file_path, num=10):
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        # 获取第一个工作表
        sheet = workbook.active
        
        # 获取列标题并确定需要移除的列
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        masked_columns = set()
        visible_headers = []
        column_titles = []  # 存储每列的原始标题
        
        # 识别需要移除的列（标题为"磁力"的列）
        for col_idx, header in enumerate(header_row):
            if header == "磁力":
                masked_columns.add(col_idx)
            else:
                header_text = format_cell(header, None)  # 格式化表头
                visible_headers.append(header_text)
                column_titles.append(header)  # 存储原始列标题
        
        # 计算每列的最大显示宽度（考虑宽窄字符）
        col_display_widths = [get_display_width(header) for header in visible_headers]
        
        # 收集所有行的可见列数据
        all_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            visible_cells = []
            for col_idx, cell in enumerate(row):
                if col_idx not in masked_columns:
                    # 使用列标题确定格式规则
                    col_title = column_titles[len(visible_cells)]
                    cell_text = format_cell(cell, col_title)
                    visible_cells.append(cell_text)
                    
                    # 更新列显示宽度
                    cell_width = get_display_width(cell_text)
                    col_index = len(visible_cells) - 1
                    if cell_width > col_display_widths[col_index]:
                        col_display_widths[col_index] = cell_width
            
            # 检查是否空行（所有可见列都为空）
            if any(cell != "" for cell in visible_cells):
                all_rows.append(visible_cells)
        
        # 计算分隔线长度
        total_display_width = sum(col_display_widths) + (len(col_display_widths) - 1) * 3 + 2
        separator = "-" * total_display_width
        
        # 打印表头
        testarr = [0,3,4,5,6]

        print(separator)
        header_parts = []
        for i, header in enumerate(visible_headers):
            padded_header = pad_to_width(header, col_display_widths[i])
            header_parts.append(padded_header+ ("" if i in testarr else "\t"))
        print("| "+"| ".join(header_parts)+"|")
        print(separator)
        
        # 打印数据行
        
        
        for row_data in random.sample(all_rows, num):
            row_parts = []
            for i, cell in enumerate(row_data):
                padded_cell = pad_to_width(cell, col_display_widths[i])
                row_parts.append(padded_cell + ("" if i in testarr else "\t"))
            print("| "+"| ".join(row_parts)+"|")
        
        print(separator)
        print(f"共读取 {sheet.max_row} 行，显示 {num} 行数据")
        
    except FileNotFoundError:
        print(f"错误: 文件 '{file_path}' 未找到")
        sys.exit(1)
    except Exception as e:
        print(f"读取文件时出错: {e}")
        sys.exit(1)

def randomlink(xlsxfile,num = 10):
    workbook = openpyxl.load_workbook(xlsxfile)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    linkarr = []
    for col_idx, header in enumerate(header_row):
        if header == "磁力":
            for row_idx,row in enumerate(sheet.iter_rows(min_row=2, values_only=True),start=1):
                link = row[col_idx]
                if link:
                    if link not in linkarr:
                        linkarr.append(link)
            break
    
    
    if len(linkarr) <= num:
        for link in linkarr:
            print(link)
    else:
        for link in random.sample(linkarr, num):
            print(link)
        

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("用法: python read_xlsx.py random/print <文件路径> <数量>")
        print("示例: python read_xlsx.py random data.xlsx 10")
        sys.exit(1)
    
    mode = sys.argv[1]
    file_path = sys.argv[2]
    num = sys.argv[3]
    if num.isdigit(): 
        num = int(num)
    else:
        print("数量参数必须是一个整数")
        sys.exit(1)

    if mode == "random":
        randomlink(file_path,num)
    elif mode == "print":
        read_xlsx(file_path,num)
    