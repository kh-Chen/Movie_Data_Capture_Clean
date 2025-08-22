import re
import os
import time
import traceback
from urllib.parse import urljoin
import json

import logger
import config
from .scraper import cover_json_data
from utils import httprequest, translate
from utils.event import register_event
from utils.number_parser import get_number
from .scrapinglib.custom.javdb import Javdb

from lxml import etree
import openpyxl

columns     = ["number","title","original_title","actor","userrating","uservotes","release","magnet_link","magnet_meta","magnet_tags",]
xlsx_hearer = ["番号",  "标题",  "原标题",        "演员", "评分",      "人数",      "发布日期","磁力",       "内容",        "标签"]
downloaded_numbers = []
catched_numbers = []
numberindex = {}

exit_now = False
def SIGINT_callback():
    global exit_now
    exit_now = True
    logger.info(f"SIGINT_callback: exit_now={exit_now}")
    
#python main.py -u https://javdb459.com/users/want_watch_videos want.xlsx
def run(arr:list):
    register_event("SIGINT", callback=SIGINT_callback)
    url = arr[0]
    xlsxfile = arr[1] if len(arr) > 1 else "scrapingurl.xlsx"
    xlsxfile = os.path.abspath(xlsxfile)

    dirs = ["/mnt/f/1","/mnt/f/downloaded","/mnt/f/store"]
    
    if not url.startswith("http"):
        domain = config.getStrValue("overGFW.javdb")
        if domain.endswith("/") and url.startswith("/"):
            url = domain[:-1] + url
        elif not domain.endswith("/") and not url.startswith("/"):
            url = domain + "/" + url
        else:
            url = domain + url

    logger.info(f'use url: {url}')

    for dir in dirs:
        filelist = os.listdir(dir)
        for file in filelist:
            if os.path.isfile(os.path.join(dir,file)):
                downloaded_numbers.append(get_number(file))

    number_index,title_index,original_title_index = 0,0,0
    for col_idx, header in enumerate(xlsx_hearer):
        if header == "番号":
            number_index = col_idx
        elif header == "标题":
            title_index = col_idx
        elif header == "原标题":
            original_title_index = col_idx

    workbook = None
    sheet = None
    if os.path.exists(xlsxfile):
        logger.info(f"file [{xlsxfile}] already exists... ")
        workbook = openpyxl.load_workbook(xlsxfile)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for col_idx, header in enumerate(xlsx_hearer):
            if header != header_row[col_idx]:
                logger.error("title not match.")
                return

        rows_to_delete = []
        rows_to_tran = []

        for row_idx,row in enumerate(sheet.iter_rows(min_row=2, values_only=True),start=2):
            number = row[number_index]
            if number is None:
                number = ''
            if number in downloaded_numbers or number == '':
                logger.info(f'will delete row: {row_idx} {row}')
                rows_to_delete.append(row_idx)
                continue
            
            catched_numbers.append(number)
            numberindex[number] = row_idx
                
            title = row[title_index]
            otitle = row[original_title_index]
            if title is None or title.strip() == '' :
                if otitle is not None and otitle.strip() != '':
                    rows_to_tran.append(row_idx)
    else:
        logger.info(f"file [{xlsxfile}] not exists, create new file... ")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"  
        sheet.append(xlsx_hearer)

    javdb(url, sheet)

    if len(rows_to_tran) > 0:
        logger.info("tran data...")
        for rowindex in rows_to_tran:
            if exit_now:
                break    
            number = sheet.cell(row=rowindex, column=number_index+1).value
            otitle = sheet.cell(row=rowindex, column=original_title_index+1).value
            zh = translate.translate_text(otitle)
            sheet.cell(row=rowindex, column=title_index+1).value = zh
            logger.info(f"{rowindex} -- {number} -- {otitle} -- {zh}")


    if len(rows_to_delete) > 0:
        logger.info(f"delete {len(rows_to_delete)} rows in xlsx file.")
        for row_idx in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_idx)
    workbook.save(xlsxfile)

#url中存在page参数时只拉取本页数据，不含page参数时则自动翻页拉取全部数据
def javdb(url:str, sheet:openpyxl.worksheet.worksheet.Worksheet) : 
    session = httprequest.request_session(cookies=Javdb.get_cookies())
    try:
        getOtherPage = 'page=' not in url
        _url = url
        pageAt = 1
        while True:
            if exit_now:
                break    
            if getOtherPage and pageAt != 1:
                url = _url + ('&' if "?" in _url else '?') + 'page=' + str(pageAt)

            resp = session.get(url)
            tree = etree.fromstring(resp.text, etree.HTMLParser()) 
            sleep(10)
            datalen = 0
            if 'want_watch_videos' in url:
                datalen = want_watch_videos(resp.url, tree, sheet, session)
                if datalen == -1:
                    return
            else:
                datalen = other(resp.url, tree, sheet, session)
                
            if not getOtherPage or datalen < 20:
                break
            pageAt += 1

        Javdb(session).save_cookies()
    except Exception as e:
        logger.error(f"url scraper error. {e}")
        logger.error(f"{traceback.format_exc()}")

def want_watch_videos(baseurl:str, tree:etree._Element, sheet:openpyxl.worksheet.worksheet.Worksheet, session):
    tag_a = tree.xpath('//*[contains(@class,"movie-list")]/div/div/a')
    datalen = len(tag_a)
    logger.info(f"get {datalen} urls in page {baseurl}")
    if datalen == 0:
        logger.info(etree.tostring(tree, encoding='unicode',pretty_print=True))


    for a in tag_a:
        if exit_now:
            break   
        detail_url,number,original_title = get_info_in_taga(a)
        if number == '':
            logger.error(f"no number in {etree.tostring(a, encoding='unicode',pretty_print=True)}")
            continue

        if number in downloaded_numbers or number in catched_numbers:
            logger.info(f"{number} already downloaded or in xlsx, skip.")
            if number in catched_numbers:
                #增量模式，出现重复数据代表增量结束。
                return -1 
            if number in numberindex:
                otitle = sheet.cell(row=numberindex[number], column=3).value
                if otitle is None:
                    otitle = ""
                if otitle != original_title:
                    sheet.cell(row=numberindex[number], column=3).value = original_title
            continue

        detail_url = urljoin(baseurl, detail_url)

        for i in range(3):
            try:
                data = get_data(detail_url,Javdb(session))
                break
            except Exception as e:
                logger.error(f"get_data from url {detail_url} error.",e)


        if data['number'] is None or data['number'] == '':
            logger.info(f"{number} error.")
            continue
        else:
            logger.info(f"{data['number']} loaded.")
        
        wdata = cover_wdata(data)
        sheet.append(wdata)
        sleep()
    
    return datalen

def get_info_in_taga(tag_a):
    detail_url = tag_a.get('href')
    tag_num = tag_a.find('div[@class="video-title"]/strong')
    number = ''
    if tag_num != None:
        number = tag_num.text
        
    original_title = ''
    tag_title = tag_a.find('div[@class="video-title"]') 
    if tag_title != None:
        text_nodes = tag_title.xpath('text()')
        original_title = ''.join(text_nodes).strip()
    return detail_url,number,original_title


def other(baseurl:str, tree:etree._Element, sheet:openpyxl.worksheet.worksheet.Worksheet, session):

    detail_urls = tree.xpath('//*[contains(@class,"movie-list")]/div/a/@href')
    if len(detail_urls) == 0 :
        detail_urls = tree.xpath('//*[contains(@class,"movie-list")]/div/div/a/@href')
    datalen = len(detail_urls)
    logger.info(f"get {datalen} urls in page {baseurl}")

    for detail_url in detail_urls:
        if exit_now:
            break    
        detail_url = urljoin(baseurl, detail_url)
        
        data = get_data(detail_url,Javdb(session))
        logger.info(f"{data['number']} loaded.")
        if data['number'] in downloaded_numbers:
            continue
        
        wdata = cover_wdata(data)
        sheet.append(wdata)
        sleep()
    
    return datalen
    

def get_data(detail_url:str,parser=None):
    json_data = parser.get_from_detail_url(detail_url)
    if not json_data:
        logger.error(f"{detail_url} load error.")
        return None
    return cover_json_data(json.loads(json_data))

def cover_wdata(data:dict):
    best = getBestMagnet(data["magnets"])
    wdata = []
    for index, key in enumerate(columns):
        if "magnet" in key:
            if "" == best:
                wdata.append("")
            else:
                if key == "magnet_link":
                    wdata.append(best["link"])
                elif key == "magnet_meta":
                    wdata.append(best["meta"])
                elif key == "magnet_tags":
                    wdata.append(",".join(best["tags"]))
        else:
            if data[key] is None or data[key] == '':
                logger.error(f"empty data.. {key} ")
            wdata.append(data[key])
    return wdata


def getBestMagnet(arr):
    if len(arr) == 0:
        return ""
    result = None
    sc = 0
    for item in arr:
        scope = 0
        if '字幕' in item["tags"]:
            scope += 100

        if "1個文件" in item["meta"]:
            scope += 4

        try:
            size = None
            _re = re.search(r'([\d\.]*)(?=GB)',item["meta"])
            if _re is None:
                _re = re.search(r'([\d\.]*)(?=MB)',item["meta"])
                if _re is not None:
                    size  = float(_re.group(0))/1000
            else:
                size  = float(_re.group(0))         

            if size is not None:
                scope += size if size < 8 else 8   
        except Exception as e:
            logger.error(f"get Magnet link size error. {item} {e}")

        if scope > sc or result is None:
            sc = scope
            result = item
    return result

def sleep(add:int=0):
    if exit_now:
        return
    interval = config.getIntValue("common.interval")
    interval += add
    if interval != 0:
        logger.info(f"Continue in {interval} seconds")
        time.sleep(interval)
    return interval